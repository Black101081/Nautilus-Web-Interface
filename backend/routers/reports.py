"""
Reports Router
Provides PDF and Excel report exports for trading performance, orders, and positions.
"""

import io
from datetime import datetime, timezone
from typing import Any, Dict, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

import database
from auth_jwt import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

async def _get_report_data() -> Dict[str, Any]:
    orders = await database.list_orders()
    positions = await database.list_db_positions(open_only=False)
    filled = [o for o in orders if o.get("status", "").lower() == "filled"]
    pnls = [float(o.get("pnl") or 0.0) for o in filled]

    total_pnl = sum(pnls)
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "orders": orders,
        "positions": positions,
        "summary": {
            "total_orders": len(orders),
            "filled_orders": len(filled),
            "total_pnl": round(total_pnl, 4),
            "gross_profit": round(sum(wins), 4),
            "gross_loss": round(sum(losses), 4),
            "win_rate": round(len(wins) / len(pnls), 4) if pnls else 0.0,
            "total_positions": len(positions),
        },
    }


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

@router.get("/excel")
async def export_excel(_user: dict = Depends(get_current_user)):
    """Export trading report as Excel (.xlsx) with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")

    data = await _get_report_data()
    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")

    ws_summary.append(["Nautilus Trader — Performance Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append(["Generated", data["generated_at"]])
    ws_summary.append([])

    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    for key, val in data["summary"].items():
        ws_summary.append([key.replace("_", " ").title(), val])

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    # ── Orders sheet ──────────────────────────────────────────────────────────
    ws_orders = wb.create_sheet("Orders")
    order_cols = ["id", "instrument", "side", "type", "quantity", "price", "status", "pnl", "timestamp"]
    ws_orders.append(order_cols)
    for cell in ws_orders[1]:
        cell.font = header_font
        cell.fill = header_fill
    for order in data["orders"]:
        ws_orders.append([order.get(c, "") for c in order_cols])
    for i, col in enumerate(order_cols, 1):
        ws_orders.column_dimensions[get_column_letter(i)].width = 18

    # ── Positions sheet ───────────────────────────────────────────────────────
    ws_pos = wb.create_sheet("Positions")
    pos_cols = ["id", "instrument", "side", "quantity", "avg_price", "unrealized_pnl", "realized_pnl", "status"]
    ws_pos.append(pos_cols)
    for cell in ws_pos[1]:
        cell.font = header_font
        cell.fill = header_fill
    for pos in data["positions"]:
        ws_pos.append([pos.get(c, "") for c in pos_cols])
    for i, col in enumerate(pos_cols, 1):
        ws_pos.column_dimensions[get_column_letter(i)].width = 18

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"nautilus_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

@router.get("/pdf")
async def export_pdf(_user: dict = Depends(get_current_user)):
    """Export trading report as PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        )
    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab not installed")

    data = await _get_report_data()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph("Nautilus Trader — Performance Report", styles["Title"]))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Generated: {data['generated_at']}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    # Summary table
    story.append(Paragraph("Summary", styles["Heading2"]))
    story.append(Spacer(1, 0.2 * cm))

    summary_rows = [["Metric", "Value"]]
    for k, v in data["summary"].items():
        summary_rows.append([k.replace("_", " ").title(), str(v)])

    tbl = Table(summary_rows, colWidths=[8 * cm, 5 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.5 * cm))

    # Recent orders table (last 20)
    recent_orders = data["orders"][-20:]
    if recent_orders:
        story.append(Paragraph("Recent Orders (last 20)", styles["Heading2"]))
        story.append(Spacer(1, 0.2 * cm))
        order_headers = ["Instrument", "Side", "Qty", "Price", "Status", "PnL"]
        order_rows = [order_headers]
        for o in recent_orders:
            order_rows.append([
                o.get("instrument", ""),
                o.get("side", ""),
                str(o.get("quantity", "")),
                str(o.get("price", "")),
                o.get("status", ""),
                str(round(float(o.get("pnl") or 0.0), 4)),
            ])
        col_w = [4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm]
        o_tbl = Table(order_rows, colWidths=col_w)
        o_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4057")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(o_tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"nautilus_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
